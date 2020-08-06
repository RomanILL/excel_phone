import openpyxl  # открывать xlsx файлы
import os
import re


def make_good_phone_list(phone_string):
    # делим строку с телефонами на отдельные телефоны
    delimiters = ".", ",", ":", ";", "\\", "/", "+", "или", "и"
    regexPattern = "|".join(map(re.escape, delimiters))
    mobile_phone_candidate_list = re.split(regexPattern, phone_string)

    # чистим от мусора и городских телефонов
    rus_standard_mobile_list = make_mobile_list(mobile_phone_candidate_list)

    return rus_standard_mobile_list


def make_mobile_list(origin_phone_list):
    good_phone_list = list()
    # перебираем номера из оригинального списка
    for phone_candidate in origin_phone_list:
        full_number = ""

        for number_x in phone_candidate:
            if number_x in "0123456789":
                full_number += number_x

        if len(full_number) == 11 and (full_number[:2] == "79" or full_number[:2] == "89"):
            full_number = make_standard_rus(full_number)
            good_phone_list.append(full_number)

        elif len(full_number) == 10 and full_number[:1] == "9":
            full_number = "7" + full_number
            full_number = make_standard_rus(full_number)
            good_phone_list.append(full_number)

    return good_phone_list


def make_standard_rus(number_11):
    """ 79xx1234567 -> +7 (9xx) 1234567 """
    number_11 = "+7 (" + number_11[1:4] + ") " + number_11[4:]
    return number_11


def create_write_file(w_filename, name_first_sheet="Лист 1"):
    """ функция создания файла записи xlsx
    возвращает объект нового файла"""
    temp_file_name = w_filename.split("\\")[-1]
    print(f'Создаем файл "{temp_file_name}"')
    # создаем новый excel-файл
    exit_file = openpyxl.Workbook()
    # добавляем новый лист
    exit_file.active.title = name_first_sheet
    return exit_file


def check_destination_folders(folders_list):
    for folder in folders_list:
        print(check_and_create_dir(folder))
    print("-" * 30 + "\n")


def make_cities_dict(file_name, country_select_list):
    # формируем словарь городов
    cities_xlsx = open_file_xlsx(file_name)
    cities_dict = dict()
    for i in range(2, cities_xlsx.active.max_row + 1):
        country_name = cities_xlsx.active.cell(row=i, column=2).value
        if country_name in country_select_list:
            city_name = cities_xlsx.active.cell(row=i, column=1).value
            region_name = cities_xlsx.active.cell(row=i, column=3).value
            cities_dict[city_name] = (country_name, region_name)
    cities_xlsx.close()
    del cities_xlsx
    return cities_dict


def make_regions_dict(file_name):
    # формируем словарь регионов
    regions_xlsx = open_file_xlsx(file_name)
    regions_dict = dict()
    for i in range(2, regions_xlsx.active.max_row + 1):
        for one_num in range(len(str(regions_xlsx.active.cell(row=i, column=1).value).split(", "))):
            region_name = regions_xlsx.active.cell(row=i, column=2).value
            regions_dict[one_num] = region_name
    regions_xlsx.close()
    del regions_xlsx
    return regions_dict


def print_any_list(any_list):
    """функция просто печатает списки в столбик"""
    count_any_list = len(any_list)
    for i in range(count_any_list):
        print(i, ") -", any_list[i])
    print("Количество элементов списка:", count_any_list)


def check_and_create_dir(dir_name):
    """функция проверяет существует ли папка, и если её нет, то создает новую"""
    if not os.path.isdir(dir_name):
        os.mkdir(dir_name)
        return f"Папка '{dir_name}' не существовала. Папка создана"
    return f"Папка '{dir_name}' существует"


def open_file_xlsx(origin_xlsx_name):
    """функция для открытия любых xlsx файлов
    возвращает кортеж: файл,
    при ошибке открытия возвращает None кортеж"""
    try:
        new_xlsx_file_object = openpyxl.load_workbook(origin_xlsx_name)
    except:
        temp_file_name = origin_xlsx_name.split("\\")[-1]
        print(f'ВНИМАНИЕ: Ошибка открытия файла "{temp_file_name}"')
        input("Чтобы продолжить нажмите Enter...")
        return None
    return new_xlsx_file_object


def get_heads(xlsx_object, sheet_number=0, start_row=1, ignore_id_list=tuple()):
    """ функция формирования списка заголовков таблицы
    на вход принимает объект открытого файла xlsx, лист, с которого считывать, номер строки, где лежат заголовки
    нумерация листов идет с 0, нумрация строк с 1"""
    heads = list()
    if len(xlsx_object.sheetnames) < sheet_number:
        print(f"Ошибка: Количество листов в xlsx книге меньше, чем {sheet_number}")
        print(f"Будет использован лист, который был активен")
    else:
        xlsx_object.active = sheet_number
    print("Для создания заголовков выбран лист:", xlsx_object.active)
    for i in range(1, xlsx_object.active.max_column + 1):
        if i not in ignore_id_list:
            head = str(xlsx_object.active.cell(row=start_row, column=i).value)
            heads.append(head)
    return heads
